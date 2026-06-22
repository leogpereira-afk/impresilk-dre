import openpyxl, json, re

SRC = '/Users/leonardopereira/Downloads/Financeiro - DRE_Caixa (Ricardo).xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['Base']

MONTHS = ['Dez/2025', 'Jan/2026', 'Fev/2026', 'Mar/2026', 'Abr/2026', 'Mai/2026']
MONTH_KEYS = ['2025-12', '2026-01', '2026-02', '2026-03', '2026-04', '2026-05']

def num(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip()
    if s == '': return 0.0
    if '.' in s and ',' in s:        # 1.234,56 -> 1234.56
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:                   # 680,00 -> 680.00
        s = s.replace(',', '.')
    try: return float(s)
    except: return 0.0

# find last account row
last = 0
for r in range(2, 1001):
    a = ws.cell(row=r, column=1).value
    if a is not None and str(a).strip():
        last = r

accounts = []
for r in range(2, last + 1):
    raw = ws.cell(row=r, column=1).value
    if raw is None: continue
    raw = str(raw).strip()
    if raw in ('Receitas', 'Despesas', 'Resultado'): continue
    m = re.match(r'^([\d\s\.]+?)\s*-\s*(.+)$', raw)
    if not m: continue
    code_part, label = m.group(1), m.group(2).strip()
    segs = [s for s in code_part.split('.') if s.strip()]
    code = '.'.join(s.strip() for s in segs)
    level = len(segs)
    vals = [round(num(ws.cell(row=r, column=2 + i).value), 2) for i in range(6)]
    parent = '.'.join(code.split('.')[:-1]) if level > 1 else None
    accounts.append({
        'code': code, 'name': label, 'level': level,
        'parent': parent, 'values': vals,
    })

out = {
    'company': 'Impresilk',
    'basis': 'Competência de Caixa',
    'months': MONTHS,
    'monthKeys': MONTH_KEYS,
    'accounts': accounts,
}
with open('data.js', 'w', encoding='utf-8') as f:
    f.write('window.DRE_DATA = ')
    json.dump(out, f, ensure_ascii=False)
    f.write(';\n')

# sanity check totals
rec = next(a for a in accounts if a['code'] == '1')
desp = next(a for a in accounts if a['code'] == '2')
print('accounts:', len(accounts))
print('Receitas:', rec['values'])
print('Despesas:', desp['values'])
print('Resultado:', [round(rec['values'][i] - desp['values'][i], 2) for i in range(6)])
